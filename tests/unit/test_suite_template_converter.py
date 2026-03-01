"""Unit tests for SuiteTemplateConverter.

All Excel workbooks are built in-memory with openpyxl so no fixture files are needed.
"""

import io
import os
import tempfile
from pathlib import Path

import openpyxl
import pytest
import yaml

from src.config.suite_template_converter import SuiteTemplateConverter
from src.contracts.test_suite import TestSuiteConfig


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_wb_bytes(tests_rows=None, config_rows=None, include_config=True):
    """
    Build an in-memory Excel workbook and return its bytes.

    tests_rows  – list of tuples for data rows on the Tests sheet.
                  If None, only the header row is written.
    config_rows – list of (key, value) tuples for the Config sheet.
                  If None and include_config is True, defaults are used.
    include_config – whether to create the Config sheet at all.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tests"

    # Header
    ws.append([
        "test_name", "test_type", "file", "mapping",
        "rules", "oracle_query", "key_columns",
        "max_errors", "max_warnings", "max_missing_rows",
        "max_extra_rows", "max_different_rows_pct",
    ])

    for row in (tests_rows or []):
        ws.append(list(row))

    if include_config:
        ws_cfg = wb.create_sheet("Config")
        ws_cfg.append(["key", "value"])
        if config_rows is not None:
            for k, v in config_rows:
                ws_cfg.append([k, v])
        else:
            ws_cfg.append(["suite_name", "My Suite"])
            ws_cfg.append(["environment", "staging"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _save_and_convert(wb_bytes, filename="suite.xlsx", output_dir=None):
    """Write bytes to a temp file, run convert(), return (yaml_path, converter)."""
    with tempfile.TemporaryDirectory() as tmp:
        xlsx_path = os.path.join(tmp, filename)
        with open(xlsx_path, "wb") as f:
            f.write(wb_bytes)

        out_dir = output_dir or tmp
        converter = SuiteTemplateConverter()
        yaml_path = converter.convert(xlsx_path, out_dir)
        with open(yaml_path, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f)
        return data, yaml_path


# ---------------------------------------------------------------------------
# Test 1: Minimal row (only required columns non-blank) → valid YAML
# ---------------------------------------------------------------------------

def test_convert_minimal_row():
    """A row with only the four required columns produces a well-formed YAML."""
    wb_bytes = _make_wb_bytes(
        tests_rows=[("P327 structural check", "structural", "data/p327.dat", "P327_mapping",
                     None, None, None, None, None, None, None, None)],
        config_rows=[("suite_name", "Minimal Suite"), ("environment", "dev")],
    )
    data, _ = _save_and_convert(wb_bytes)

    assert data["name"] == "Minimal Suite"
    assert data["environment"] == "dev"
    assert len(data["tests"]) == 1
    t = data["tests"][0]
    assert t["name"] == "P327 structural check"
    assert t["type"] == "structural"
    assert t["file"] == "data/p327.dat"
    assert t["mapping"] == "P327_mapping"
    assert "thresholds" in t
    assert t["thresholds"]["max_errors"] == 0  # default


# ---------------------------------------------------------------------------
# Test 2: Row with all optional columns → correct threshold values
# ---------------------------------------------------------------------------

def test_convert_row_with_all_optional_columns():
    """All optional threshold columns are included in the output."""
    wb_bytes = _make_wb_bytes(
        tests_rows=[(
            "Full test", "oracle_vs_file", "data/full.dat", "FULL_MAP",
            "rules/r.json", "SELECT 1 FROM DUAL", "id,name",
            5, 10, 2, 3, 1.5,
        )],
        config_rows=[("suite_name", "Full Suite"), ("environment", "prod")],
    )
    data, _ = _save_and_convert(wb_bytes)

    t = data["tests"][0]
    assert t["rules"] == "rules/r.json"
    assert t["oracle_query"] == "SELECT 1 FROM DUAL"
    assert t["key_columns"] == ["id", "name"]
    assert t["thresholds"]["max_errors"] == 5
    assert t["thresholds"]["max_warnings"] == 10
    assert t["thresholds"]["max_missing_rows"] == 2
    assert t["thresholds"]["max_extra_rows"] == 3
    assert t["thresholds"]["max_different_rows_pct"] == 1.5


# ---------------------------------------------------------------------------
# Test 3: Config sheet sets suite_name and environment
# ---------------------------------------------------------------------------

def test_config_sheet_sets_suite_name_and_environment():
    wb_bytes = _make_wb_bytes(
        tests_rows=[],
        config_rows=[("suite_name", "P327 UAT Suite"), ("environment", "prod")],
    )
    data, _ = _save_and_convert(wb_bytes)

    assert data["name"] == "P327 UAT Suite"
    assert data["environment"] == "prod"


# ---------------------------------------------------------------------------
# Test 4: Missing Config sheet → suite_name inferred from filename
# ---------------------------------------------------------------------------

def test_missing_config_sheet_infers_suite_name():
    wb_bytes = _make_wb_bytes(tests_rows=[], include_config=False)
    data, _ = _save_and_convert(wb_bytes, filename="my_uat_suite.xlsx")

    assert data["name"] == "my_uat_suite"
    assert data["environment"] == "dev"  # default


# ---------------------------------------------------------------------------
# Test 5: Blank optional cells → fields omitted from YAML (not null)
# ---------------------------------------------------------------------------

def test_blank_optional_cells_omitted():
    """Optional fields that are blank (None) must not appear in the YAML output."""
    wb_bytes = _make_wb_bytes(
        tests_rows=[("Omit test", "structural", "f.dat", "MAP",
                     None, None, None, None, None, None, None, None)],
        config_rows=[("suite_name", "Omit Suite"), ("environment", "dev")],
    )
    data, _ = _save_and_convert(wb_bytes)

    t = data["tests"][0]
    assert "rules" not in t
    assert "oracle_query" not in t
    assert "key_columns" not in t
    # Only max_errors should be in thresholds (defaulted to 0)
    assert list(t["thresholds"].keys()) == ["max_errors"]


# ---------------------------------------------------------------------------
# Test 6: key_columns comma-separated → list in YAML
# ---------------------------------------------------------------------------

def test_key_columns_comma_separated_to_list():
    wb_bytes = _make_wb_bytes(
        tests_rows=[("Keys test", "structural", "f.dat", "MAP",
                     None, None, "col_a, col_b, col_c",
                     None, None, None, None, None)],
        config_rows=[("suite_name", "Keys Suite"), ("environment", "dev")],
    )
    data, _ = _save_and_convert(wb_bytes)

    assert data["tests"][0]["key_columns"] == ["col_a", "col_b", "col_c"]


# ---------------------------------------------------------------------------
# Test 7: Template creation – file created and has "Tests" sheet
# ---------------------------------------------------------------------------

def test_create_template_has_tests_sheet():
    converter = SuiteTemplateConverter()
    with tempfile.TemporaryDirectory() as tmp:
        tpl_path = os.path.join(tmp, "template.xlsx")
        result = converter.create_template(tpl_path)

        assert result == tpl_path
        assert os.path.exists(tpl_path)

        wb = openpyxl.load_workbook(tpl_path)
        assert "Tests" in wb.sheetnames


# ---------------------------------------------------------------------------
# Test 8: Template creation – "Tests" sheet has all expected column headers
# ---------------------------------------------------------------------------

def test_create_template_tests_sheet_has_all_headers():
    expected_headers = [
        "test_name", "test_type", "file", "mapping",
        "rules", "oracle_query", "key_columns",
        "max_errors", "max_warnings", "max_missing_rows",
        "max_extra_rows", "max_different_rows_pct",
    ]
    converter = SuiteTemplateConverter()
    with tempfile.TemporaryDirectory() as tmp:
        tpl_path = os.path.join(tmp, "template.xlsx")
        converter.create_template(tpl_path)

        wb = openpyxl.load_workbook(tpl_path)
        ws = wb["Tests"]
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert header == expected_headers


# ---------------------------------------------------------------------------
# Test 9: Template creation – has "Config" sheet with key/value columns
# ---------------------------------------------------------------------------

def test_create_template_has_config_sheet():
    converter = SuiteTemplateConverter()
    with tempfile.TemporaryDirectory() as tmp:
        tpl_path = os.path.join(tmp, "template.xlsx")
        converter.create_template(tpl_path)

        wb = openpyxl.load_workbook(tpl_path)
        assert "Config" in wb.sheetnames

        ws_cfg = wb["Config"]
        header = [cell.value for cell in next(ws_cfg.iter_rows(min_row=1, max_row=1))]
        assert header == ["key", "value"]


# ---------------------------------------------------------------------------
# Test 10: Empty Excel (header only, no data rows) → YAML with empty tests list
# ---------------------------------------------------------------------------

def test_empty_excel_produces_empty_tests_list():
    wb_bytes = _make_wb_bytes(
        tests_rows=[],
        config_rows=[("suite_name", "Empty Suite"), ("environment", "dev")],
    )
    data, _ = _save_and_convert(wb_bytes)

    assert data["tests"] == []


# ---------------------------------------------------------------------------
# Test 11: Skips rows where test_name is blank
# ---------------------------------------------------------------------------

def test_skips_rows_with_blank_test_name():
    wb_bytes = _make_wb_bytes(
        tests_rows=[
            ("Valid test", "structural", "data/a.dat", "MAP_A",
             None, None, None, None, None, None, None, None),
            (None, "structural", "data/b.dat", "MAP_B",
             None, None, None, None, None, None, None, None),
            ("  ", "structural", "data/c.dat", "MAP_C",
             None, None, None, None, None, None, None, None),
        ],
        config_rows=[("suite_name", "Skip Suite"), ("environment", "dev")],
    )
    data, _ = _save_and_convert(wb_bytes)

    # Only "Valid test" row should survive
    assert len(data["tests"]) == 1
    assert data["tests"][0]["name"] == "Valid test"


# ---------------------------------------------------------------------------
# Test 12: max_different_rows_pct string "5.0" → float 5.0
# ---------------------------------------------------------------------------

def test_max_different_rows_pct_string_converts_to_float():
    """Even if the cell contains a string value it must be stored as a float."""
    # Build workbook manually so we can insert a string into the numeric cell
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tests"
    ws.append([
        "test_name", "test_type", "file", "mapping",
        "rules", "oracle_query", "key_columns",
        "max_errors", "max_warnings", "max_missing_rows",
        "max_extra_rows", "max_different_rows_pct",
    ])
    # Insert "5.0" as a string in the last column
    ws.append(["Pct test", "structural", "f.dat", "MAP",
               None, None, None, None, None, None, None, "5.0"])

    ws_cfg = wb.create_sheet("Config")
    ws_cfg.append(["key", "value"])
    ws_cfg.append(["suite_name", "Pct Suite"])
    ws_cfg.append(["environment", "dev"])

    buf = io.BytesIO()
    wb.save(buf)
    wb_bytes = buf.getvalue()

    data, _ = _save_and_convert(wb_bytes)

    assert data["tests"][0]["thresholds"]["max_different_rows_pct"] == 5.0
    assert isinstance(data["tests"][0]["thresholds"]["max_different_rows_pct"], float)


# ---------------------------------------------------------------------------
# Test 13: Roundtrip – converted YAML validates against TestSuiteConfig model
# ---------------------------------------------------------------------------

def test_roundtrip_validates_against_pydantic_model():
    """The generated YAML must load cleanly into the TestSuiteConfig Pydantic model."""
    wb_bytes = _make_wb_bytes(
        tests_rows=[
            ("Structural check", "structural", "data/p327.dat", "P327_full",
             None, None, None, 0, None, None, None, None),
            ("Oracle comparison", "oracle_vs_file", "data/target.dat", "TGT_MAP",
             None, "SELECT * FROM TGT", "id,name",
             0, 5, 2, 3, 2.5),
            ("Rules check", "rules", "data/src.dat", "SRC_MAP",
             "rules/biz_rules.json", None, None,
             0, 10, None, None, None),
        ],
        config_rows=[("suite_name", "Roundtrip Suite"), ("environment", "staging")],
    )
    data, _ = _save_and_convert(wb_bytes)

    # This must not raise
    suite = TestSuiteConfig(**data)
    assert suite.name == "Roundtrip Suite"
    assert suite.environment == "staging"
    assert len(suite.tests) == 3
    assert suite.tests[1].key_columns == ["id", "name"]
    assert suite.tests[2].rules == "rules/biz_rules.json"
