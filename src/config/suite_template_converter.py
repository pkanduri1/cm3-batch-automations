"""Convert Excel test suite templates to YAML configuration files."""

import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
import yaml


# Columns that map to threshold fields in the output YAML
_THRESHOLD_COLUMNS = {
    "max_errors",
    "max_warnings",
    "max_missing_rows",
    "max_extra_rows",
    "max_different_rows_pct",
}

# All recognised column names on the Tests sheet
_TEST_COLUMNS = [
    "test_name",
    "test_type",
    "file",
    "mapping",
    "rules",
    "oracle_query",
    "key_columns",
    "max_errors",
    "max_warnings",
    "max_missing_rows",
    "max_extra_rows",
    "max_different_rows_pct",
]

# Columns on the Config sheet
_CONFIG_COLUMNS = ["key", "value"]


def _sanitize_filename(name: str) -> str:
    """Replace characters that are unsafe in filenames with underscores."""
    return re.sub(r"[^\w\-.]", "_", name)


def _cell_value(cell_val: Any) -> Any:
    """Return the cell value, stripping whitespace from strings."""
    if cell_val is None:
        return None
    if isinstance(cell_val, str):
        stripped = cell_val.strip()
        return stripped if stripped != "" else None
    return cell_val


class SuiteTemplateConverter:
    """Convert an Excel test-suite template to a YAML file consumed by cm3-batch run-tests."""

    # ------------------------------------------------------------------ #
    # Public API                                                           #
    # ------------------------------------------------------------------ #

    def convert(self, excel_path: str, output_dir: str = ".") -> str:
        """
        Read an Excel file and produce a YAML test suite file.

        Returns the path to the generated YAML file.
        """
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

        suite_name, environment = self._read_config_sheet(wb, excel_path)
        tests = self._read_tests_sheet(wb)

        wb.close()

        suite_data: Dict[str, Any] = {
            "name": suite_name,
            "environment": environment,
            "tests": tests,
        }

        os.makedirs(output_dir, exist_ok=True)
        safe_name = _sanitize_filename(suite_name)
        output_path = os.path.join(output_dir, f"{safe_name}.yaml")

        with open(output_path, "w", encoding="utf-8") as f:
            yaml.dump(suite_data, f, default_flow_style=False, allow_unicode=True)

        return output_path

    def create_template(self, output_path: str) -> str:
        """
        Write an empty Excel template that users can fill in.

        Returns output_path.
        """
        wb = openpyxl.Workbook()

        # ---- Tests sheet ------------------------------------------------
        ws_tests = wb.active
        ws_tests.title = "Tests"
        ws_tests.append(_TEST_COLUMNS)

        # ---- Config sheet -----------------------------------------------
        ws_config = wb.create_sheet("Config")
        ws_config.append(_CONFIG_COLUMNS)
        ws_config.append(["suite_name", "My Test Suite"])
        ws_config.append(["environment", "dev"])

        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        wb.save(output_path)
        return output_path

    # ------------------------------------------------------------------ #
    # Private helpers                                                      #
    # ------------------------------------------------------------------ #

    def _read_config_sheet(self, wb: openpyxl.Workbook, excel_path: str):
        """Return (suite_name, environment) from the Config sheet, or infer from filename."""
        suite_name = Path(excel_path).stem
        environment = "dev"

        if "Config" not in wb.sheetnames:
            return suite_name, environment

        ws = wb["Config"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return suite_name, environment

        # Find header row to determine column order
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        try:
            key_idx = header.index("key")
            val_idx = header.index("value")
        except ValueError:
            return suite_name, environment

        config: Dict[str, str] = {}
        for row in rows[1:]:
            if len(row) <= max(key_idx, val_idx):
                continue
            k = _cell_value(row[key_idx])
            v = _cell_value(row[val_idx])
            if k is not None and v is not None:
                config[str(k).strip().lower()] = str(v).strip()

        if "suite_name" in config:
            suite_name = config["suite_name"]
        if "environment" in config:
            environment = config["environment"]

        return suite_name, environment

    def _read_tests_sheet(self, wb: openpyxl.Workbook) -> List[Dict[str, Any]]:
        """Parse the Tests sheet and return a list of test dicts."""
        if "Tests" not in wb.sheetnames:
            return []

        ws = wb["Tests"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        # Header row
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]

        tests: List[Dict[str, Any]] = []
        for row in rows[1:]:
            test = self._parse_test_row(header, row)
            if test is not None:
                tests.append(test)

        return tests

    def _parse_test_row(self, header: List[str], row: tuple) -> Optional[Dict[str, Any]]:
        """Convert a single spreadsheet row to a test dict, or None if it should be skipped."""

        def get(col: str) -> Any:
            """Return the value for a named column, or None if absent / blank."""
            try:
                idx = header.index(col)
            except ValueError:
                return None
            val = _cell_value(row[idx]) if idx < len(row) else None
            return val

        # Skip rows with a blank test_name
        test_name = get("test_name")
        if test_name is None:
            return None

        test: Dict[str, Any] = {
            "name": str(test_name).strip(),
            "type": str(get("test_type") or "structural").strip(),
            "file": str(get("file") or "").strip(),
            "mapping": str(get("mapping") or "").strip(),
        }

        # Optional scalar fields
        rules_val = get("rules")
        if rules_val is not None:
            test["rules"] = str(rules_val).strip()

        oracle_query_val = get("oracle_query")
        if oracle_query_val is not None:
            test["oracle_query"] = str(oracle_query_val).strip()

        key_columns_val = get("key_columns")
        if key_columns_val is not None:
            test["key_columns"] = [c.strip() for c in str(key_columns_val).split(",") if c.strip()]

        # Threshold fields
        thresholds: Dict[str, Any] = {}

        max_errors_val = get("max_errors")
        if max_errors_val is not None:
            thresholds["max_errors"] = int(max_errors_val)
        else:
            # max_errors defaults to 0 per the TestSuiteConfig model
            thresholds["max_errors"] = 0

        for col in ("max_warnings", "max_missing_rows", "max_extra_rows"):
            val = get(col)
            if val is not None:
                thresholds[col] = int(val)

        pct_val = get("max_different_rows_pct")
        if pct_val is not None:
            thresholds["max_different_rows_pct"] = float(pct_val)

        test["thresholds"] = thresholds

        return test
